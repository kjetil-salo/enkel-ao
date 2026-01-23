
import openpyxl
import json
from pathlib import Path

# Finn prosjektroten (der denne fila ligger)
ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / 'docs' / 'Norgeslisten_pr_31-12-2024_AviList_v2025.xlsx'
JSON_PATH = ROOT / 'public' / 'data' / 'norske_arter.json'

# Åpne arbeidsbok og første ark
wb = openpyxl.load_workbook(XLSX_PATH, read_only=True)
ws = wb.active

species = []

# Finn kolonneindekser for "Norsk navn" og "Vitenskapelig navn"
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
col_norsk = None
col_latin = None
for i, col in enumerate(header_row):
    if str(col).strip().lower() == 'norsk navn':
        col_norsk = i
    if str(col).strip().lower() == 'vitenskapelig navn':
        col_latin = i

species = []
for row in ws.iter_rows(min_row=2, values_only=True):
    name = row[col_norsk] if col_norsk is not None else None
    latin = row[col_latin] if col_latin is not None else None
    if not name or not latin:
        continue
    # Hopp over rader som kun er overskrifter eller grupper
    if str(name).strip() == '' or str(latin).strip() == '':
        continue
    # Kun arter/underarter: latin-navn må ha minst to ord (dvs. ikke "Anseriformes", "Anatidae" osv.)
    if len(str(latin).strip().split()) < 2:
        continue
    species.append({
        'norwegian': name,
        'latin': latin
    })

with open(JSON_PATH, 'w', encoding='utf-8') as f:
    json.dump(species, f, ensure_ascii=False, indent=2)

print(f'Lagret {len(species)} arter til {JSON_PATH}')
